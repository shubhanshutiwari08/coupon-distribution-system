from flask import Flask, render_template, request
from openpyxl import Workbook
from openpyxl import load_workbook
import sqlite3
import pandas as pd
# import random
# import string
import qrcode
from io import BytesIO
import base64



app = Flask(__name__)

# FOR FORM DATA
def init_db():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS form_data
                (id INTEGER PRIMARY KEY AUTOINCREMENT,
                 name TEXT,
                 cpfno TEXT,
                 email TEXT,
                 num_coupons INTEGER)''')
    conn.commit()
    conn.close()

# Call the init_db function to create the database and table
init_db()

# FOR COUPON DATA
def init_db_coupon():
    conn = sqlite3.connect('coupon_database.db')
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS coupon_data
                (id INTEGER PRIMARY KEY AUTOINCREMENT,
                 cpfno TEXT,
                 num_coupons INTEGER,
                 num_coupons_used INTEGER)''')
    conn.commit()
    conn.close()

# Call the init_db function to create the database and table
init_db_coupon()


# TO WRITE INTO EXCEL for FORM
# def write_to_excel(field1_value, field2_value, field3_value, field4_value):
#     try:
#         workbook = load_workbook('formdata.xlsx')
#     except FileNotFoundError:
#         workbook = Workbook()
    
#     # Select the active sheet
#     sheet = workbook.active
    
#     # Find the first empty row
#     row = sheet.max_row + 1
    
#     # Write the input values to the first and second columns of the empty row
#     sheet.cell(row=row, column=1).value = field1_value
#     sheet.cell(row=row, column=2).value = field2_value
#     sheet.cell(row=row, column=3).value = field3_value
#     sheet.cell(row=row, column=4).value = field4_value
    
#     # Save the workbook
#     workbook.save('formdata.xlsx')


# TO WRITE INTO EXCEL for COUPON
def write_to_excel_coupon(field1_value, field2_value, field3_value):
    try:
        workbook = load_workbook('coupondata.xlsx')
    except FileNotFoundError:
        workbook = Workbook()
    
    # Select the active sheet
    sheet = workbook.active
    
    # Find the first empty row
    row = sheet.max_row + 1
    
    # Write the input values to the first and second columns of the empty row
    sheet.cell(row=row, column=1).value = field1_value
    sheet.cell(row=row, column=2).value = field2_value
    sheet.cell(row=row, column=3).value = field3_value
    
    # Save the workbook
    workbook.save('coupondata.xlsx')


# VENDER
@app.route('/vender', methods=['GET', 'POST'])
def vender():
    if request.method == 'POST':
        cpfno = request.form['cpfno']
        num_coupons = request.form['num_coupons']
        num_coupons_used = request.form['num_coupons_used']

        # n = int(num_coupons)
        a = int(cpfno)
        b = int(num_coupons)
        c = int(num_coupons_used)

        # Read the Excel sheet data
        excel_data = pd.read_excel('empdata.xlsx')

        # Check if the data already exists in the Excel sheet
        # matching_data = excel_data[(excel_data['cpfno'] == cpfno) & (excel_data['email'] == email)]
        matching_data = excel_data[(excel_data['cpfno'] == a) & (b <= excel_data['num_coupons']) & (c <= b) ]

        if  matching_data.empty:
            return render_template('invalid.html')
        else:
            conn = sqlite3.connect('coupon_database.db')
            c = conn.cursor()
            # c.execute('SELECT * FROM form_data WHERE cpfno=? AND email=?', (cpfno, email))
            c.execute('SELECT * FROM coupon_data WHERE cpfno=?', (cpfno,))
            stored_data = c.fetchone()
            conn.close()

            if stored_data:
                return render_template('already_filled.html')
            else:
                # Add the data to the database
                conn = sqlite3.connect('coupon_database.db')
                c = conn.cursor()
                c.execute('INSERT INTO coupon_data (cpfno, num_coupons, num_coupons_used) VALUES (?, ?, ?)', (cpfno, num_coupons, num_coupons_used))

            conn.commit()
            # conn.close()

            # coupons = generate_coupons(n)
            # Get the input values from the form
            # field1_input = request.form['field1']
            # field2_input = request.form['field2']

            # Call the function to write the values to an Excel file
            write_to_excel_coupon(cpfno, num_coupons, num_coupons_used)

            return render_template('coupon_active.html', cpfno=cpfno, num_coupons=num_coupons, num_coupons_used=num_coupons_used )



    return render_template('vender.html')


@app.route('/coupondata')
def coupon_data():
    conn = sqlite3.connect('coupon_database.db')
    c = conn.cursor()
    c.execute('SELECT * FROM coupon_data')
    data = c.fetchall()

    # Find duplicate entries based on name and email
    duplicates = set()
    unique_data = []
    for row in data:
        if (row[1], row[2]) in duplicates:
            duplicates.add((row[1], row[2]))
        else:
            unique_data.append(row)

    conn.close()

    return render_template('coupon_data.html', data=unique_data, duplicates=duplicates)
# VENDER ENDS



# Route to handle form submission

# Route to handle form submission
@app.route('/')
def main():
    return render_template("home.html")


@app.route('/coupon', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # name = request.form['name']
        cpfno = request.form['cpfno']
        # email = request.form['email']
        num_coupons = request.form['num_coupons']

        c = int(cpfno)
        d = int(num_coupons)
        

        # Read the Excel sheet data
        excel_data = pd.read_excel('empdata.xlsx')

        # Check if the data already exists in the Excel sheet
        # matching_data = excel_data[(excel_data['cpfno'] == cpfno) & (excel_data['email'] == email)]
        matching_data = excel_data[(excel_data['cpfno'] == c) & (d <= excel_data['num_coupons'])]

        if  matching_data.empty:
            return render_template('invalid.html')
        else:
            conn = sqlite3.connect('database.db')
            c = conn.cursor()
            c.execute('SELECT * FROM form_data WHERE cpfno=?', (cpfno,))

            stored_data = c.fetchone()
            conn.close()

            if stored_data:
                return render_template('already_filled.html')
            else:
                # Add the data to the database
                conn = sqlite3.connect('database.db')
                c = conn.cursor()
                c.execute('INSERT INTO form_data (cpfno, num_coupons) VALUES (?, ?)', (cpfno, num_coupons))

            conn.commit()
            # conn.close()

            # coupons = generate_coupons(n)
                      
                        
            qr_data = f"CPF Number: {cpfno}\n Number of Coupons: {num_coupons}"
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(qr_data)
            qr.make(fit=True)
            qr_img = qr.make_image(fill="black", back_color="white")

    # Save the QR code to a BytesIO buffer
            qr_buffer = BytesIO()
            qr_img.save(qr_buffer)
            qr_image_data = qr_buffer.getvalue()
    
    
    
            
            # write_to_excel(name, email, cpfno, num_coupons)
            # return render_template('success.html', coupons=coupons,cpfno=cpfno )
            return render_template('success.html', qr_image_data=base64.b64encode(qr_image_data).decode('utf-8'), cpfno=cpfno)



    return render_template('index.html')


#admin

@app.route('/admin')
def admin_access():
    return render_template('admin.html')



    # Route to display the database content
@app.route('/data')
def data():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('SELECT * FROM form_data')
    data = c.fetchall()

    # Find duplicate entries based on name and email
    duplicates = set()
    unique_data = []
    for row in data:
        if ( row[1], row[2], row[3]) in duplicates:
            duplicates.add(( row[1],row[2],row[3]))
        else:
            unique_data.append(row)

    conn.close()

    return render_template('data.html', data=unique_data, duplicates=duplicates)




# Route to generate and store the number of coupons
# @app.route('/generate_coupon', methods=['GET','POST'])
# def generate_coupon():
#     if request.method == 'POST':
#         num_coupons = request.form['num_coupons']

#     conn = sqlite3.connect('database.db')
#     c = conn.cursor()

#     # Get the total number of rows in the form_data table
#     # c.execute('SELECT COUNT(*) FROM form_data')
#     # total_rows = c.fetchone()[0]

#     # Check if the total number of rows has reached the limit (1500)
#     # if total_rows >= 1500:
#     #     return render_template('coupon_limit.html')

#     # Store the number of coupons in the database
#     c.execute('INSERT INTO form_data (num_coupons) VALUES (?)', (num_coupons))
#     conn.commit()
#     conn.close()

#     return render_template('data.html')




# Route to clear the database
@app.route('/clear', methods=['GET'])
def clear_data():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('DELETE FROM form_data')
    conn.commit()
    conn.close()
    return "Database cleared successfully"

@app.route('/final', methods=['GET','POST'])
def final_coupon():
    # if request.method == 'POST':
    #     coupon = request.form['coupon']
    return render_template("final.html")

# Run the Flask application
if __name__ == '__main__':
    app.run(port=2020)




